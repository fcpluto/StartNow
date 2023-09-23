## Python读取不同格式文件

# 1.文本文件 (.txt)
with open('file.txt','r') as file:
    content = file.read()

#2.CSV文件 (.csv)
import csv

with open('data.csv','r') as file:
      csv_reader = csv.reader(file)
      for row in csv_reader:
        print(row)

#3.Excel文件 (.xlsx)
from openpyxl import load_workbook
workbook = load_workbook('data.xlsx')
sheet = workbook.active
for row in sheet.iter_rows(values_only=True):
    print(row)


# 4.JSON文件 (json)
import json
with open('data.json','r') as file:
    data = json.load(file)

# 5.XML文件 (.xml)
import xml.etree.ElementTree as ET
tree = ET.parse('data.xml')
root = tree.getroot()

# 6.SQLite数据库文件 (.sglite)
import sqlite3
conn = sqlite3.connect('data.sqlite')
cursor = conn.cursor()
cursor.execute('SELECT * FROM table_name')


# 7.PDF文件 (.pdf)
import PyPDF2
with open('document.pdf','rb') as file:
    pdf_reader = PyPDF2.PdfFileReader(file)
    text = ''
    for page_num in range(pdf_reader.numPages):
        text += pdf_reader.getPage(page_num).extractText()


# 8.Word文件 (docx)
from docx import Document
doc = Document('document.docx')
text = ''
for paragraph in doc.paragraphs:
    text += paragraph.text
